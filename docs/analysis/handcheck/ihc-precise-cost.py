#!/usr/bin/env python3
"""
状态：历史手核输入(2026-07-06)。本脚本=一抗试剂层手核(¥38/例量级·标准成本口径·非全成本)，是 golden-registry G-COST-3 的复现脚本（勿删/勿归档=会孤儿化黄金锚）。脚本内“真精算/真价”指逐抗体台账价 vs 均价法的内部对比、非“真实成本”（当前口径权威 P0 spec：“真”只修饰实收与片数）。当前成本口径权威见 docs/COREONE-成本口径-P0内圈-院级贡献毛利-绝对最小业务逻辑-2026-07-04.md。

免疫组化逐抗体真精算(一抗试剂层) — 2026-07-01
数据(PII安全,只用这两个;不用含患者隐私的"病理样本信息汇总"):
  DET = ~/Downloads/0702免组.xlsx        逐抗体行(病理号/蜡块/切片/markerName/adviceType)
  TAI = ~/Downloads/免疫组化相关耗材2025年.xlsx  sheet'2025 (2)'=抗体台账
口径:
  真抗体 = adviceType in (Y000001,Y000003) [排除 Y000006 HE/深切重切, Y000007 白片]
  每人份成本 = 台账第14列"单价"(已换算,如2SC即用¥99.82) ; 兜底=瓶价(col5)/换算率(col11)
  抗体名归一化 = 去括号克隆号+去空格/连字符+大写 (Ki67↔Ki-67, PD-L1(22C3)↔PD-L1)
注意: 台账第0行是标题、第1行才是列名; 有两列都叫"单价"(col5瓶价/col13每人份价) — 取col13直接用,勿再除换算率(初版bug)。
仅一抗层; "算全"还需 显色试剂盒+工时+设备(走G2)。14未命中=特染混入(PAS/GMS)+台账缺的抗体。
用法: python3 ihc-precise-cost.py <0702免组.xlsx> <免疫组化耗材2025.xlsx>
"""
import openpyxl, sys, collections, re, statistics

def norm(s):
    s = str(s or '').upper().strip()
    s = re.sub(r'[\(（].*?[\)）]', '', s)   # 去括号克隆号
    return s.replace(' ', '').replace('-', '').replace('_', '')

def main(det_path, tai_path):
    # 台账 → 抗体库 {norm(name): {name,type,per(每人份)}}
    ws = openpyxl.load_workbook(tai_path, read_only=True, data_only=True)['2025 (2)']
    tr = list(ws.iter_rows(min_row=1, values_only=True))
    lib = {}
    for r in tr[2:]:
        if not r[0]:
            continue
        def num(i):
            try:
                return float(r[i]) if i < len(r) and r[i] not in (None, '') else 0
            except Exception:
                return 0
        per = num(13) or (num(5) / num(11) if num(11) > 0 else 0)
        if per <= 0:
            continue
        lib[norm(r[0])] = {'name': str(r[0]).strip(), 'type': r[2], 'per': per}
    print("台账抗体品种(有效):", len(lib))

    # LIS 明细 → 每例真抗体
    ws2 = openpyxl.load_workbook(det_path, read_only=True, data_only=True).worksheets[0]
    dr = list(ws2.iter_rows(min_row=1, values_only=True))
    dh = {x: i for i, x in enumerate(dr[0])}
    def D(row, n):
        i = dh.get(n)
        return row[i] if i is not None and i < len(row) else None
    cases = collections.defaultdict(list)
    for r in dr[1:]:
        if str(D(r, 'adviceType')) in ('Y000001', 'Y000003'):
            cases[str(D(r, 'caseNo'))].append(str(D(r, 'markerName')).strip())

    allab = [m for v in cases.values() for m in v]
    matched = [m for m in allab if norm(m) in lib]
    unmatched = sorted(set(m for m in allab if norm(m) not in lib))
    print("真抗体条数 %d 命中 %d = %d%%" % (len(allab), len(matched), len(matched) / len(allab) * 100))
    print("未命中(%d): %s" % (len(unmatched), unmatched))

    percase = {cn: sum(lib[norm(m)]['per'] for m in mks if norm(m) in lib) for cn, mks in cases.items()}
    costs = list(percase.values())
    print("每例一抗试剂成本: 病例%d Σ¥%.0f 中位¥%.0f 均值¥%.0f max¥%.0f"
          % (len(percase), sum(costs), statistics.median(costs), statistics.mean(costs), max(costs)))
    avg = sum(lib[norm(m)]['per'] for m in matched) / len(matched)
    err = [abs(avg * sum(1 for m in mks if norm(m) in lib) - percase[cn]) / percase[cn]
           for cn, mks in cases.items() if percase[cn] > 0]
    print("均价法 ¥%.1f/抗体 vs 逐抗体真价: 每例误差 中位%d%% max%d%%"
          % (avg, statistics.median(err) * 100, max(err) * 100))

if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
